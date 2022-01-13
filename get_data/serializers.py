from rest_framework import serializers
from rest_framework.validators import ValidationError
from .models import *
from product.tasks import create_or_update_products, get_order, get_ozon_transaction, get_analitic_data
from datetime import datetime, date
from datetime import timedelta
import openpyxl


class UserSerializer(serializers.ModelSerializer):

    """
    Пользователь

    Сериализатор для создания пользователя, изменения пароля или данных для подключения к ozon и в целом пользователя.

    Поля которые необходимо указывать при сохранении данных маркетплейса
        marketplace_id
        api_key
        marketplace_name

    Поля которые необходимо указывать при изменении пароля:
        new_password

    Для изменения данны
    """

    password = serializers.CharField(required=False)
    email = serializers.EmailField(required=False)

    def create(self, validated_data):
        user = User(**validated_data)
        password = validated_data.get("password", None)
        user.set_password(password)
        user.save()
        return user

    def update(self, instance, validated_data):
        user_id = self.context['request'].user

        if validated_data.get("api_key") is not None:
            marketplace_id = str(validated_data['marketplace_id'])
            api_key = validated_data['api_key']
            marketplace_name = validated_data['marketplace_name']
            pk = self.context['pk']
            user = User.objects.get(pk=pk)

            if marketplace_name == "ozon":
                api_key_isset = requests.post('https://api-seller.ozon.ru/v1/product/list',  headers={'Client-Id': marketplace_id,
                                                                                                      'Api-Key': api_key,
                                                                                                      'Content-Type': 'application/json',
                                                                                                      'Host': 'api-seller.ozon.ru'})
                if api_key_isset.status_code == 200:
                    last_validations_date = datetime.now()
                    marketplace = Marketplace.objects.create_marketplace(marketplace_name=marketplace_name, marketplace_id=marketplace_id,
                                                                         api_key=api_key, last_validations_date=last_validations_date)

                    user.marketplace_data.add(marketplace.pk)

                    # instance.save()
                    # get_product.delay(user_id=user_id.id)
                    # get_order.delay(user_id=user_id.id)
                    # get_ozon_transaction.delay(user_id=user_id.id)
                    # get_analitic_data.delay(user_id=user_id.id)
                    return marketplace
                else:
                    raise ValidationError(
                        detail={"Invalid Api-Key, please contact support": "404"}
                    )

        elif validated_data.get("new_password") is not None:
            for attr, value in validated_data.items():
                setattr(instance, attr, value)

            password = validated_data.get("new_password", None)
            instance.set_password(password)

            instance.save()
            return instance

        else:
            user = User.objects.get(id=user_id.id)

            for attr, value in validated_data.items():
                setattr(instance, attr, value)

            password = validated_data.get("password", None)

            if user.password == password:
                instance.save()
                return instance
            elif password == None:
                instance.save()
                return instance

    new_password = serializers.CharField(max_length=32, write_only=True, required=False)
    marketplace_id = serializers.IntegerField(write_only=True, required=False)
    marketplace_name = serializers.CharField(max_length=32, write_only=True, required=False)


    class Meta:
        model = User
        fields = ("id", "email", "password", "first_name", "last_name", "patronymic", "role", "date_create",
                  "post_agreement", 'card', "card_year", "card_ovner", "ozon_id", "api_key", 'name_org', 'bank', 'inn',
                  'orgn', 'kpp', 'bank_account', 'correspondent_bank_account', 'bik', 'new_password', 'user_tarif_data',
                  'return_status', 'marketplace_id', 'marketplace_name')


class TransactionSerializer(serializers.ModelSerializer):
    """
        Транзакции внутри приложения
    """

    def create(self, validated_data):

        # Получаем существующие данные из exel

        user_tariff_registration = openpyxl.open("users.xlsx", read_only=True)
        sheet = user_tariff_registration.active
        list_array = []

        for row in range(1, sheet.max_row + 1):
            fio = sheet[row][0].value
            email = sheet[row][1].value
            rate = sheet[row][2].value
            list_array.append({
                "fio": fio,
                "email": email,
                "rate": rate,
            })

        # Сохраняем в excel файл

        write_user_tariff_registration = openpyxl.Workbook()
        sheet = write_user_tariff_registration.active

        if not list_array:
            sheet['A1'] = "ФИО"
            sheet['B1'] = "E-mail"
            sheet['C1'] = "Тариф"

        first_name = validated_data.get("id_user").first_name
        last_name = validated_data.get("id_user").last_name
        patronymic = validated_data.get("id_user").patronymic

        email = validated_data.get("id_user").email

        if first_name or last_name or patronymic is not None:
            fio = f'{first_name} {last_name} {patronymic}'
        else:
            fio = 'Данные пользователя в личном кабинете не заполненены'

        rate = f'{validated_data.get("rate").rate_name}'
        list_array.append({
            "fio": fio,
            "email": email,
            "rate": rate,
        })

        row = 1
        for user in list_array:
            sheet.cell(row=row, column=1).value = user['fio']
            sheet.cell(row=row, column=2).value = user['email']
            sheet.cell(row=row, column=3).value = user['rate']
            row += 1

        write_user_tariff_registration.save("users.xlsx")
        write_user_tariff_registration.close()

        # Создание транзакции
        transaction = Transaction(**validated_data)

        rate_pk = validated_data.get("rate").pk
        user_pk = validated_data.get("id_user").pk
        rate = Rate.objects.get(pk=rate_pk)
        user = User.objects.get(pk=user_pk)

        if validated_data.get("summ") == rate.price:
            transaction.save()
            user.role = 'subscription'
            user.save()
            task_date = datetime.utcnow() + timedelta(days=rate.validity)
            create_or_update_products.apply_async(id_user=user_pk, eta=task_date)

        return transaction

    class Meta:

        model = Transaction
        fields = "__all__"


class RateSerializer(serializers.ModelSerializer):
    """
        Тариф
    """

    class Meta:
        model = Rate
        fields = "__all__"
