from .models import Transaction
from ..account.models import User
from ..rate.models import  Rate
from rest_framework import serializers
from ..account.tasks import create_or_update_products
from datetime import datetime, date
from datetime import timedelta
import openpyxl

class TransactionSerializer(serializers.ModelSerializer):
    """
        Транзакции внутри приложения
    """

    def create(self, validated_data):

        # Получаем существующие данные из exel

        user_tariff_registration = openpyxl.open("users.xlsx", read_only=True)
        sheet = user_tariff_registration.active
        list_array = []

        for row in range(1, sheet.max_row + 1):
            fio = sheet[row][0].value
            email = sheet[row][1].value
            rate = sheet[row][2].value
            list_array.append({
                "fio": fio,
                "email": email,
                "rate": rate,
            })

        # Сохраняем в excel файл

        write_user_tariff_registration = openpyxl.Workbook()
        sheet = write_user_tariff_registration.active

        if not list_array:
            sheet['A1'] = "ФИО"
            sheet['B1'] = "E-mail"
            sheet['C1'] = "Тариф"

        first_name = validated_data.get("id_user").first_name
        last_name = validated_data.get("id_user").last_name
        patronymic = validated_data.get("id_user").patronymic

        email = validated_data.get("id_user").email

        if first_name or last_name or patronymic is not None:
            fio = f'{first_name} {last_name} {patronymic}'
        else:
            fio = 'Данные пользователя в личном кабинете не заполненены'

        rate = f'{validated_data.get("rate").rate_name}'
        list_array.append({
            "fio": fio,
            "email": email,
            "rate": rate,
        })

        row = 1
        for user in list_array:
            sheet.cell(row=row, column=1).value = user['fio']
            sheet.cell(row=row, column=2).value = user['email']
            sheet.cell(row=row, column=3).value = user['rate']
            row += 1

        write_user_tariff_registration.save("users.xlsx")
        write_user_tariff_registration.close()

        # Создание транзакции
        transaction = Transaction(**validated_data)

        rate_pk = validated_data.get("rate").pk
        user_pk = validated_data.get("id_user").pk
        rate = Rate.objects.get(pk=rate_pk)
        user = User.objects.get(pk=user_pk)

        if validated_data.get("summ") == rate.price:
            transaction.save()
            user.role = 'subscription'
            user.save()
            task_date = datetime.utcnow() + timedelta(days=rate.validity)
            create_or_update_products.apply_async(id_user=user_pk, eta=task_date)

        return transaction

    class Meta:

        model = Transaction
        fields = "__all__"